#!/usr/bin/env python3
"""
Teaching Assistant Brain Generator (v2 — Excel-only input)
===========================================================
Reads a course planning Excel + student roster and produces
a complete brain/ folder ready for any LLM assistant.

No config.yaml needed. Everything comes from the Excel sheets.

Usage:
    python generate_brain.py <course_folder>

    where <course_folder> contains:
        inputs/
            course_plan.xlsx   (the planning Excel with all sheets)
            roster.xlsx        (student roster — optional)

    Output is written to <course_folder>/brain/
    Canvas pages are written to <course_folder>/canvas_pages/

Dependencies:
    pip install openpyxl
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
TA_ROOT = SCRIPT_DIR.parent
TEMPLATES_DIR = TA_ROOT / "templates"
SHARED_DIR = TA_ROOT / "shared"


# ---------------------------------------------------------------------------
# Excel → Config: read all sheets into a unified config dict
# ---------------------------------------------------------------------------

def read_excel_config(excel_path: Path) -> dict:
    """Read the course planning Excel and return a config dict."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    config = {}

    # ── Sheet: Course Info (key-value pairs) ──
    if "Course Info" in wb.sheetnames:
        ws = wb["Course Info"]
        kv = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and str(row[0]).strip():
                key = str(row[0]).strip()
                val = row[1]
                if val is not None:
                    kv[key] = val

        config["course_name"] = str(kv.get("Course Name", "Untitled Course"))
        config["course_code"] = str(kv.get("Course Code", "COURSE"))
        config["quarter"] = str(kv.get("Quarter", ""))
        config["university"] = str(kv.get("University", ""))
        config["instructor"] = str(kv.get("Instructor", ""))
        config["student_id_prefix"] = str(kv.get("Student ID Prefix", "s"))

        config["schedule"] = {
            "days": [d.strip() for d in str(kv.get("Teaching Days", "")).split(",") if d.strip()],
            "first_class": _format_date(kv.get("First Class Date")),
            "last_class": _format_date(kv.get("Last Class Date")),
            "total_classes": _safe_int(kv.get("Total Classes", 20)),
            "class_duration_min": _safe_int(kv.get("Class Duration (min)", 100)),
        }

        config["quizzes"] = {
            "total": _safe_int(kv.get("Total Quizzes", 0)),
            "drop_lowest": _safe_int(kv.get("Drop Lowest", 0)),
        }

        mid_class = kv.get("Midterm Class Number")
        if mid_class:
            config["midterm"] = {
                "class_number": _safe_int(mid_class),
                "format": str(kv.get("Midterm Format", "")),
                "weight": str(kv.get("Midterm Weight", "")),
            }

        proj_name = kv.get("Project Name")
        if proj_name:
            config["final_project"] = {
                "name": str(proj_name),
                "due_class": _safe_int(kv.get("Project Due Class", 0)),
                "weight": str(kv.get("Project Weight", "")),
            }

        config["features"] = {
            "case_based": _is_yes(kv.get("Case-Based Course")),
            "guest_speakers": _is_yes(kv.get("Guest Speakers")),
            "group_presentations": _is_yes(kv.get("Group Presentations")),
            "strategy_in_news": _is_yes(kv.get("Strategy in News")),
        }

    # ── Sheet: Sections ──
    config["sections"] = []
    if "Sections" in wb.sheetnames:
        ws = wb["Sections"]
        headers = _get_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            name = _get_val(vals, headers, "Section Name")
            if not name:
                continue
            config["sections"].append({
                "name": name,
                "days": _get_val(vals, headers, "Days") or "",
                "time": _get_val(vals, headers, "Time") or "",
                "enrollment": _safe_int(_get_val(vals, headers, "Enrollment") or 0),
            })

    if not config["sections"]:
        config["sections"] = [{"name": "Section 1", "days": "", "time": "", "enrollment": 0}]

    # ── Sheet: Sessions (the big planning sheet) ──
    config["sessions"] = []
    if "Sessions" in wb.sheetnames:
        ws = wb["Sessions"]
        headers = _get_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            sno = _get_val(vals, headers, "Session #")
            if sno is None:
                continue
            try:
                sno = int(sno)
            except (ValueError, TypeError):
                continue

            sess = {
                "number": sno,
                "date": _format_date(_get_val(vals, headers, "Date")),
                "day": _expand_day(_get_val(vals, headers, "Day") or ""),
                "week_number": _safe_int(_get_val(vals, headers, "Week #") or 0),
                "week_theme": _get_val(vals, headers, "Week Theme"),
                "topic": _get_val(vals, headers, "Topic") or "",
                "session_type": _get_val(vals, headers, "Session Type") or "lecture",
                "key_question": _get_val(vals, headers, "Key Question"),
                "required_reading": _get_val(vals, headers, "Required Reading"),
                "case_name": _get_val(vals, headers, "Case Name"),
                "optional_reading": _get_val(vals, headers, "Optional Reading"),
                "assignment_due": _get_val(vals, headers, "Assignment Due"),
                "activity_name": _get_val(vals, headers, "Activity Name"),
                "activity_description": _get_val(vals, headers, "Activity Description"),
                "canvas_intro": _get_val(vals, headers, "Canvas: Session Intro"),
                "canvas_prep": _get_val(vals, headers, "Canvas: Prep Instructions"),
                "canvas_in_class": _get_val(vals, headers, "Canvas: In Class"),
                "canvas_after_class": _get_val(vals, headers, "Canvas: After Class"),
                "canvas_prep_questions": _get_val(vals, headers, "Canvas: Prep Questions"),
                "notes": _get_val(vals, headers, "Notes / Thoughts"),
                # Dynamic course management columns
                "status": _get_val(vals, headers, "Status") or "planned",
                "revised_topic": _get_val(vals, headers, "Revised Topic"),
                "revised_reading": _get_val(vals, headers, "Revised Reading"),
                # AI pedagogy columns
                "ai_mode": _get_val(vals, headers, "AI Mode"),
                "ai_role": _get_val(vals, headers, "AI Role"),
                "ai_activity": _get_val(vals, headers, "AI Activity"),
            }
            config["sessions"].append(sess)

    # ── Sheet: Change Log ──
    config["change_log"] = []
    if "Change Log" in wb.sheetnames:
        ws = wb["Change Log"]
        headers = _get_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            change_date = _get_val(vals, headers, "Date")
            if not change_date:
                continue
            config["change_log"].append({
                "date": _format_date(change_date),
                "session": _safe_int(_get_val(vals, headers, "Session #") or 0),
                "change_type": _get_val(vals, headers, "Change Type") or "",
                "what_changed": _get_val(vals, headers, "What Changed") or "",
                "why": _get_val(vals, headers, "Why") or "",
                "moved_where": _get_val(vals, headers, "What Moved Where") or "",
                "impact": _get_val(vals, headers, "Impact on Other Sessions") or "",
            })

    # ── Sheet: Grading ──
    config["grading"] = {}
    if "Grading" in wb.sheetnames:
        ws = wb["Grading"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and row[1]:
                config["grading"][str(row[0]).strip()] = str(row[1]).strip()

    # ── Sheet: Guest Speakers ──
    config["guest_speakers"] = []
    if "Guest Speakers" in wb.sheetnames:
        ws = wb["Guest Speakers"]
        headers = _get_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            name = _get_val(vals, headers, "Speaker Name")
            if not name:
                continue
            config["guest_speakers"].append({
                "name": name,
                "session": _safe_int(_get_val(vals, headers, "Session #") or 0),
                "date": _format_date(_get_val(vals, headers, "Date")),
                "topic": _get_val(vals, headers, "Topic / Title") or "",
                "notes": _get_val(vals, headers, "Bio / Notes") or "",
            })

    # ── Sheet: SET Baseline ──
    config["set_baseline"] = {}
    if "SET Baseline" in wb.sheetnames:
        ws = wb["SET Baseline"]
        headers = _get_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            q = _get_val(vals, headers, "Question")
            label = _get_val(vals, headers, "Label")
            baseline = _get_val(vals, headers, "Your Baseline")
            target = _get_val(vals, headers, "Target")
            notes = _get_val(vals, headers, "Notes")
            if q:
                config["set_baseline"][q] = {
                    "label": label or "",
                    "baseline": baseline,
                    "target": target,
                    "notes": notes or "",
                }

    # ── Sheet: Class Template ──
    config["class_template"] = []
    if "Class Template" in wb.sheetnames:
        ws = wb["Class Template"]
        headers = _get_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            seg = _get_val(vals, headers, "Segment")
            mins = _get_val(vals, headers, "Minutes")
            if seg and mins:
                config["class_template"].append({
                    "segment": seg,
                    "minutes": _safe_int(mins),
                })

    if not config["class_template"]:
        config["class_template"] = [
            {"segment": "Attendance + Agenda", "minutes": 5},
            {"segment": "Hook / Recap", "minutes": 10},
            {"segment": "Concept 1", "minutes": 20},
            {"segment": "Activity 1", "minutes": 20},
            {"segment": "Break", "minutes": 5},
            {"segment": "Concept 2 / Discussion", "minutes": 20},
            {"segment": "Activity 2", "minutes": 10},
            {"segment": "Recap + Preview", "minutes": 10},
        ]

    wb.close()
    return config


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _get_headers(ws):
    """Get column headers from the first row as a dict of name→index."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1
    return headers


def _get_val(values, headers, key):
    """Safely get a value from a row using the header map."""
    idx = headers.get(key)
    if idx is None or idx >= len(values):
        return None
    val = values[idx]
    if val is None:
        return None
    return str(val).strip() if not isinstance(val, (int, float, datetime, date)) else val


def _format_date(val):
    """Format a date value to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _safe_int(val):
    """Safely convert to int."""
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _is_yes(val):
    """Check if a value means 'yes'."""
    if val is None:
        return False
    return str(val).strip().lower() in ("yes", "y", "true", "1")


def _expand_day(abbrev):
    """Expand day abbreviation to full name."""
    day_map = {
        "T": "Tuesday", "TH": "Thursday", "M": "Monday",
        "W": "Wednesday", "F": "Friday",
        "Tue": "Tuesday", "Thu": "Thursday", "Mon": "Monday",
        "Wed": "Wednesday", "Fri": "Friday",
        "Tu": "Tuesday", "Th": "Thursday",
    }
    return day_map.get(abbrev.strip(), abbrev.strip())


# ---------------------------------------------------------------------------
# Roster parsing
# ---------------------------------------------------------------------------

def parse_roster(roster_path: Path, prefix: str) -> list[dict]:
    """Parse student roster Excel into a list of student dicts with assigned IDs."""
    wb = openpyxl.load_workbook(roster_path, data_only=True)
    ws = wb.active
    headers = _get_headers(ws)

    # Try common column name variants
    fname_keys = ["First Name", "First", "FirstName", "first_name", "Given Name"]
    lname_keys = ["Last Name", "Last", "LastName", "last_name", "Surname", "Family Name"]
    email_keys = ["Email", "E-mail", "email", "Email Address"]

    fname_col = None
    for k in fname_keys:
        if k in headers:
            fname_col = k
            break

    lname_col = None
    for k in lname_keys:
        if k in headers:
            lname_col = k
            break

    email_col = None
    for k in email_keys:
        if k in headers:
            email_col = k
            break

    if fname_col is None:
        print(f"  WARNING: Could not find first name column in roster. Tried: {fname_keys}")
        wb.close()
        return []

    students = []
    idx = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        first = _get_val(vals, headers, fname_col)
        if not first:
            continue

        last = _get_val(vals, headers, lname_col) if lname_col else ""
        email = _get_val(vals, headers, email_col) if email_col else ""

        students.append({
            "id": f"{prefix}_{idx:03d}",
            "first_name": str(first),
            "last_name": str(last) if last else "",
            "email": str(email) if email else "",
        })
        idx += 1

    wb.close()
    return students


# ---------------------------------------------------------------------------
# Tracker JSON generation
# ---------------------------------------------------------------------------

def generate_tracker(config: dict) -> dict:
    """Generate the course tracker JSON from config."""
    schedule = config["schedule"]
    sessions = config.get("sessions", [])

    # Build sections
    sections = {}
    for s in config["sections"]:
        sections[s["name"]] = {
            "time": s.get("time", ""),
            "students": s.get("enrollment", 0),
        }

    # Build classes list
    classes = []
    for sess in sessions:
        cls = {
            "number": sess["number"],
            "date": sess.get("date"),
            "day": sess.get("day", ""),
            "topic": sess.get("topic", ""),
            "session_type": sess.get("session_type", "lecture"),
            "plan_status": sess.get("status", "planned"),
            "status": "upcoming",
        }
        if sess.get("case_name"):
            cls["case"] = sess["case_name"]
        if sess.get("key_question"):
            cls["key_question"] = sess["key_question"]
        # Track revisions
        if sess.get("revised_topic"):
            cls["revised_topic"] = sess["revised_topic"]
        if sess.get("revised_reading"):
            cls["revised_reading"] = sess["revised_reading"]
        # AI pedagogy
        if sess.get("ai_mode"):
            cls["ai_mode"] = sess["ai_mode"]
        if sess.get("ai_role"):
            cls["ai_role"] = sess["ai_role"]
        if sess.get("ai_activity"):
            cls["ai_activity"] = sess["ai_activity"]
        classes.append(cls)

    # Build quiz schedule from case sessions
    quiz_schedule = []
    quiz_num = 1
    max_quizzes = config.get("quizzes", {}).get("total", 0)
    for sess in sessions:
        if sess.get("case_name") and quiz_num <= max_quizzes:
            quiz_schedule.append({
                "quiz": quiz_num,
                "class": sess["number"],
                "date": sess.get("date"),
                "case": sess.get("case_name", ""),
                "graded": False,
            })
            quiz_num += 1

    quiz_tracker = {
        "total_quizzes": config.get("quizzes", {}).get("total", 0),
        "drop_lowest": config.get("quizzes", {}).get("drop_lowest", 0),
        "quiz_schedule": quiz_schedule,
    }

    tracker = {
        "course": config["course_name"],
        "quarter": config["quarter"],
        "sections": sections,
        "schedule": {
            "days": schedule["days"],
            "first_class": schedule["first_class"],
            "last_class": schedule["last_class"],
            "total_classes": schedule["total_classes"],
        },
        "master_input": {
            "file": "inputs/course_plan.xlsx",
        },
        "classes": classes,
        "quiz_tracker": quiz_tracker,
    }

    # Midterm
    if config.get("midterm"):
        mid = config["midterm"]
        mid_date = None
        for cls in classes:
            if cls["number"] == mid["class_number"]:
                mid_date = cls.get("date")
                break
        tracker["midterm"] = {
            "class_number": mid["class_number"],
            "date": mid_date,
            "format": mid.get("format", ""),
            "weight": mid.get("weight", ""),
        }

    # Final project
    if config.get("final_project"):
        fp = config["final_project"]
        tracker["final_project"] = {
            "name": fp.get("name", "Final Project"),
            "due_class": fp.get("due_class"),
            "weight": fp.get("weight", ""),
        }

    # Guest speakers
    if config.get("guest_speakers"):
        tracker["guest_speakers"] = config["guest_speakers"]

    # Change log
    if config.get("change_log"):
        tracker["change_log"] = config["change_log"]
    else:
        tracker["change_log"] = []

    # Action items
    tracker["action_items"] = []
    tracker["next_action_id"] = 1

    return tracker


# ---------------------------------------------------------------------------
# ASSISTANT.md generation
# ---------------------------------------------------------------------------

def generate_assistant_md(config: dict) -> str:
    """Generate the ASSISTANT.md operating manual."""
    code = config["course_code"]
    schedule = config["schedule"]
    days_str = " / ".join(schedule["days"])
    duration = schedule.get("class_duration_min", 100)
    prefix = config["student_id_prefix"]
    multi_section = len(config["sections"]) > 1

    # Section rows
    section_rows = ""
    for s in config["sections"]:
        section_rows += f"| {s['name']} | {s.get('days', days_str)}, {s.get('time', '')} (~{s.get('enrollment', '?')} students) |\n"

    # Midterm row
    midterm_row = ""
    if config.get("midterm"):
        mid = config["midterm"]
        mid_date = ""
        for sess in config.get("sessions", []):
            if sess["number"] == mid["class_number"]:
                mid_date = sess.get("date", "")
                break
        midterm_row = f"| Midterm | Class {mid['class_number']} ({mid_date}) |\n"

    # Final project row
    fp_row = ""
    if config.get("final_project"):
        fp = config["final_project"]
        fp_row = f"| {fp.get('name', 'Final Project')} | Due Class {fp.get('due_class', '?')} |\n"

    # Section terminology
    section_note = ""
    if multi_section:
        section_note = "- **Section** = a group of students. Prep is done ONCE per class and taught to all sections. Post-class notes may vary by section.\n"

    # Class template
    template_lines = ""
    for seg in config.get("class_template", []):
        template_lines += f"- {seg['segment']}: {seg['minutes']} min\n"

    # Read the template file and do replacements
    template_path = TEMPLATES_DIR / "ASSISTANT.md.template"
    with open(template_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Simple placeholder replacements
    replacements = {
        "{{course_code}}": code,
        "{{course_name}}": config["course_name"],
        "{{quarter}}": config["quarter"],
        "{{university}}": config.get("university", ""),
        "{{instructor}}": config.get("instructor", ""),
        "{{class_duration_min}}": str(duration),
        "{{total_classes}}": str(schedule["total_classes"]),
        "{{first_class}}": schedule.get("first_class", ""),
        "{{last_class}}": schedule.get("last_class", ""),
        "{{student_id_prefix}}": prefix,
    }
    for placeholder, value in replacements.items():
        content = content.replace(placeholder, value)

    # Replace course basics table
    basics_table = f"""| Field | Value |
|---|---|
| Course | {config['course_name']} |
| Quarter | {config['quarter']} |
| University | {config.get('university', '')} |
| Instructor | {config.get('instructor', '')} |
{section_rows}| Class duration | {duration} minutes |
| Total classes | {schedule['total_classes']} |
| First class | {schedule.get('first_class', '')} |
| Last class | {schedule.get('last_class', '')} |
{midterm_row}{fp_row}"""

    content = re.sub(
        r'\| Field \| Value \|.*?(?=\n---)',
        basics_table.strip(),
        content, flags=re.DOTALL
    )

    # Handle conditional sections
    content = re.sub(
        r'{{#multi_section}}.*?{{/multi_section}}',
        section_note.strip() if multi_section else "",
        content, flags=re.DOTALL
    )
    content = re.sub(
        r'{{#class_template}}.*?{{/class_template}}',
        template_lines.strip(),
        content, flags=re.DOTALL
    )
    content = re.sub(r'{{#\w+}}|{{/\w+}}', '', content)

    return content


# ---------------------------------------------------------------------------
# Memory.md generation
# ---------------------------------------------------------------------------

def generate_memory_md(config: dict) -> str:
    """Generate initial memory.md."""
    schedule = config["schedule"]
    days_str = " / ".join(schedule["days"])

    sections_lines = ""
    for s in config["sections"]:
        sections_lines += f"- {s['name']}: {s.get('days', days_str)} {s.get('time', '')}, ~{s.get('enrollment', '?')} students\n"

    grading_lines = ""
    for comp, wt in config.get("grading", {}).items():
        grading_lines += f"- {comp}: {wt}\n"

    midterm_line = ""
    if config.get("midterm"):
        midterm_line = f"- Midterm: Class {config['midterm']['class_number']}\n"

    quiz_line = ""
    q = config.get("quizzes", {})
    if q.get("total", 0) > 0:
        quiz_line = f"- Quizzes: {q['total']} total, drop lowest {q.get('drop_lowest', 0)}\n"

    return f"""# Memory — {config['course_code']} {config['quarter']}

> Rolling memory file. Updated at the end of each working session. Keep under 100 lines.

---

## Course identity

- {config['course_name']}, {config['quarter']}, {config.get('university', '')}
{sections_lines}- {schedule['total_classes']} classes, {schedule.get('class_duration_min', 100)} min each, {schedule.get('first_class', '')} to {schedule.get('last_class', '')}
{midterm_line}{quiz_line}
## Grading weights

{grading_lines if grading_lines else "- (fill in from syllabus)"}

## Current focus

- (fill in after first class)

## Patterns in this cohort

- (fill in as patterns emerge)

## Content notes

- (fill in after first few classes)

## Things to remember for next quarter

- (accumulate throughout the quarter)
"""


# ---------------------------------------------------------------------------
# Class Prep Reference generation
# ---------------------------------------------------------------------------

def generate_class_prep_reference(config: dict) -> str:
    """Generate class_prep_reference.md from sessions data."""
    sessions = config.get("sessions", [])
    lines = [
        "# Class Prep Reference\n",
        "> Generated from the course planning Excel. The tracker JSON holds live state;",
        "> this file holds the per-class detail.\n",
        "---\n",
    ]

    for sess in sessions:
        status = sess.get("status", "planned")
        status_badge = ""
        if status == "cancelled":
            status_badge = " [CANCELLED]"
        elif status == "revised":
            status_badge = " [REVISED]"
        elif status == "swapped":
            status_badge = " [SWAPPED]"

        lines.append(f"## Class {sess['number']} — {sess.get('date', 'TBD')} ({sess.get('day', '')}){status_badge}")

        # Show revised topic if it exists, with original for reference
        if sess.get("revised_topic"):
            lines.append(f"**Topic:** {sess['revised_topic']}")
            lines.append(f"~~Original topic: {sess.get('topic', 'TBD')}~~\n")
        else:
            lines.append(f"**Topic:** {sess.get('topic', 'TBD')}")
        lines.append(f"**Type:** {sess.get('session_type', 'lecture')}\n")

        if sess.get("key_question"):
            lines.append(f"**Key Question:** {sess['key_question']}\n")

        if sess.get("case_name"):
            lines.append(f"**Case:** {sess['case_name']}\n")

        # Show revised reading if it exists, with original for reference
        if sess.get("revised_reading"):
            lines.append(f"**Required Reading (revised):**\n{sess['revised_reading']}\n")
            if sess.get("required_reading"):
                lines.append(f"~~Original: {sess['required_reading']}~~\n")
        elif sess.get("required_reading"):
            lines.append(f"**Required Reading:**\n{sess['required_reading']}\n")

        if sess.get("optional_reading"):
            lines.append(f"**Optional Reading:** {sess['optional_reading']}\n")

        if sess.get("assignment_due"):
            lines.append(f"**Assignment Due:** {sess['assignment_due']}\n")

        if sess.get("activity_name"):
            lines.append(f"**Activity:** {sess['activity_name']}")
            if sess.get("activity_description"):
                lines.append(f"{sess['activity_description']}\n")
            else:
                lines.append("")

        # AI pedagogy
        if sess.get("ai_mode"):
            ai_badge = {"Solo": "🧠 Solo", "AI-Partnered": "🤖 AI-Partnered", "Reflection": "🪞 Reflection"}.get(
                sess["ai_mode"], sess["ai_mode"])
            lines.append(f"**AI Mode:** {ai_badge}")
            if sess.get("ai_role"):
                lines.append(f"**AI Role:** {sess['ai_role']} (Mollick framework)")
            if sess.get("ai_activity"):
                lines.append(f"**AI Activity:** {sess['ai_activity']}")
            lines.append("")

        if sess.get("notes"):
            lines.append(f"**Instructor Notes:** {sess['notes']}\n")

        lines.append("---\n")

    # Append change log summary if changes exist
    change_log = config.get("change_log", [])
    if change_log:
        lines.append("\n## Change Log\n")
        lines.append("> Changes made to the original plan during the quarter.\n")
        for change in change_log:
            lines.append(f"**{change.get('date', '')} — Session {change.get('session', '?')}** ({change.get('change_type', '')})")
            lines.append(f"- What: {change.get('what_changed', '')}")
            lines.append(f"- Why: {change.get('why', '')}")
            if change.get("moved_where"):
                lines.append(f"- Moved to: {change['moved_where']}")
            if change.get("impact"):
                lines.append(f"- Impact: {change['impact']}")
            lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# SET Protocol generation (with actual scores)
# ---------------------------------------------------------------------------

def generate_set_protocol(config: dict) -> str:
    """Generate set_protocol.md with actual baseline scores filled in."""
    shared_path = SHARED_DIR / "set_protocol.md"
    with open(shared_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Replace the Historical Scores table with actual data
    baseline = config.get("set_baseline", {})
    if baseline:
        scores_table = "| Question | Label | Baseline | Target | Notes |\n"
        scores_table += "|----------|-------|----------|--------|-------|\n"
        for q_key, data in baseline.items():
            b = data.get("baseline")
            t = data.get("target")
            b_str = f"{b:.2f}" if isinstance(b, (int, float)) else "—"
            t_str = f"{t:.2f}" if isinstance(t, (int, float)) else "—"
            scores_table += f"| {q_key} | {data.get('label', '')} | {b_str} | {t_str} | {data.get('notes', '')} |\n"

        content = re.sub(
            r'\| Question \| Baseline.*?(?=\n\n|\Z)',
            scores_table.strip(),
            content, flags=re.DOTALL
        )

    return content


# ---------------------------------------------------------------------------
# Canvas Pages generation
# ---------------------------------------------------------------------------

def generate_canvas_pages(config: dict, canvas_dir: Path):
    """Generate Canvas/LMS pages from session data."""
    sessions = config.get("sessions", [])
    if not sessions:
        return

    sessions_dir = canvas_dir / "Sessions"
    weeks_dir = canvas_dir / "Weeks"
    sessions_dir.mkdir(parents=True, exist_ok=True)
    weeks_dir.mkdir(parents=True, exist_ok=True)

    # ── Session pages ──
    for sess in sessions:
        date_str = sess.get("date", "TBD")
        if date_str and date_str != "TBD":
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                date_display = dt.strftime("%b %-d")
            except ValueError:
                date_display = date_str
        else:
            date_display = "TBD"

        topic = sess.get("topic", "TBD")
        filename = f"Session {sess['number']:02d} - {date_display} - {_safe_filename(topic)}.md"

        lines = [f"# Session {sess['number']} — {topic}\n"]

        if sess.get("key_question"):
            lines.append(f"**Key question:** {sess['key_question']}\n")

        if sess.get("canvas_intro"):
            lines.append(f"{sess['canvas_intro']}\n")

        # AI mode banner (right after intro, before prep)
        if sess.get("ai_mode"):
            ai_labels = {
                "Solo": "**🧠 Today's AI Mode: Solo** — Work without AI tools today. Build your own frameworks and reasoning first.",
                "AI-Partnered": "**🤖 Today's AI Mode: AI-Partnered** — Use AI as a thinking partner today. Details below.",
                "Reflection": "**🪞 Today's AI Mode: Reflection** — Reflect on your learning process — with and without AI.",
            }
            lines.append(ai_labels.get(sess["ai_mode"], f"**AI Mode: {sess['ai_mode']}**"))
            if sess.get("ai_activity"):
                lines.append(f"\n*{sess['ai_activity']}*")
            if sess.get("ai_role"):
                lines.append(f"\n*AI Role: {sess['ai_role']}*")
            lines.append("")

        lines.append("---\n")

        # Prepare section
        has_prep = sess.get("canvas_prep") or sess.get("required_reading") or sess.get("case_name")
        if has_prep:
            lines.append("## Prepare\n")
            if sess.get("canvas_prep"):
                for item in sess["canvas_prep"].split("\n"):
                    item = item.strip()
                    if item:
                        if not item.startswith("-"):
                            item = f"- {item}"
                        lines.append(item)
            elif sess.get("required_reading"):
                for item in str(sess["required_reading"]).split("\n"):
                    item = item.strip()
                    if item:
                        if not item.startswith("-"):
                            item = f"- {item}"
                        lines.append(item)
            lines.append("")

        # Prep questions
        if sess.get("canvas_prep_questions"):
            lines.append("**Preparation questions:**\n")
            for i, q in enumerate(sess["canvas_prep_questions"].split("\n"), 1):
                q = q.strip()
                if q:
                    # Remove leading number if already present
                    q = re.sub(r'^\d+\.\s*', '', q)
                    lines.append(f"{i}. {q}")
            lines.append("")

        # In Class section
        if sess.get("canvas_in_class"):
            lines.append("## In Class\n")
            for item in sess["canvas_in_class"].split("\n"):
                item = item.strip()
                if item:
                    if not item.startswith("-"):
                        item = f"- {item}"
                    lines.append(item)
            lines.append("")

        # Due this class
        if sess.get("assignment_due"):
            lines.append("**Due this class:**\n")
            for item in str(sess["assignment_due"]).split("\n"):
                item = item.strip()
                if item:
                    if not item.startswith("-"):
                        item = f"- {item}"
                    lines.append(item)
            lines.append("")

        # After Class section
        if sess.get("canvas_after_class"):
            lines.append("## After Class\n")
            lines.append(sess["canvas_after_class"])
            lines.append("")

        with open(sessions_dir / filename, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    # ── Week overview pages ──
    weeks = {}
    for sess in sessions:
        wk = sess.get("week_number", 0)
        if wk not in weeks:
            weeks[wk] = {"theme": sess.get("week_theme"), "sessions": []}
        if sess.get("week_theme"):
            weeks[wk]["theme"] = sess["week_theme"]
        weeks[wk]["sessions"].append(sess)

    for wk_num in sorted(weeks.keys()):
        if wk_num == 0:
            continue
        wk = weeks[wk_num]
        theme = wk.get("theme") or wk["sessions"][0].get("topic", "")
        filename = f"Week {wk_num:02d} - {_safe_filename(theme)}.md"

        lines = [f"# Week {wk_num} — {theme}\n"]
        lines.append("---\n")
        lines.append("## This Week\n")

        for sess in wk["sessions"]:
            day = sess.get("day", "")
            date_str = sess.get("date", "TBD")
            if date_str and date_str != "TBD":
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    date_display = dt.strftime("%b %-d")
                except ValueError:
                    date_display = date_str
            else:
                date_display = "TBD"

            lines.append(f"### {day}, {date_display} — {sess.get('topic', 'TBD')}\n")

            if sess.get("canvas_prep") or sess.get("required_reading"):
                prep_text = sess.get("canvas_prep") or str(sess.get("required_reading", ""))
                first_line = prep_text.split("\n")[0].strip()
                if first_line.startswith("- "):
                    first_line = first_line[2:]
                lines.append(f"**Prepare:** {first_line}\n")

        with open(weeks_dir / filename, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    # ── Course Schedule page ──
    lines = [f"# Course Schedule\n"]
    lines.append(f"**{config['course_name']} — {config['quarter']}**\n")
    lines.append("| Wk | Session | Date | Day | Topic |")
    lines.append("|---|---|---|---|---|")

    for sess in sessions:
        date_str = sess.get("date", "TBD")
        if date_str and date_str != "TBD":
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                date_display = dt.strftime("%b %-d")
            except ValueError:
                date_display = date_str
        else:
            date_display = "TBD"
        day_short = {"Tuesday": "Tue", "Thursday": "Thu", "Monday": "Mon",
                     "Wednesday": "Wed", "Friday": "Fri"}.get(sess.get("day", ""), sess.get("day", ""))
        lines.append(f"| {sess.get('week_number', '')} | {sess['number']} | {date_display} | {day_short} | {sess.get('topic', '')} |")

    with open(canvas_dir / "Course Schedule.md", "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # ── Canvas Setup Guide ──
    guide_lines = [
        f"# Canvas Setup Guide — {config['course_code']} {config['quarter']}\n",
        "This document tells you how to set up your Canvas course using the generated pages.\n",
        "---\n",
        "## Generated Pages\n",
        "### Session Pages (one per class)\n",
        f"Located in `Sessions/`. {len(sessions)} session pages generated.\n",
        "### Week Overview Pages\n",
        f"Located in `Weeks/`. {len(weeks) - (1 if 0 in weeks else 0)} week pages generated.\n",
        "### Course Schedule\n",
        "Located at `Course Schedule.md`. Master schedule table.\n",
        "---\n",
        "## How to Add Pages to Canvas\n",
        "1. Create one module per week (Week 1, Week 2, ...)\n",
        "2. In each module, add the week overview page first, then the two session pages\n",
        "3. In the Canvas page editor, paste the markdown content (or convert to HTML)\n",
        "4. **Publish week by week** — keep future weeks unpublished\n",
    ]

    with open(canvas_dir / "CANVAS SETUP GUIDE.md", "w", encoding="utf-8") as f:
        f.write("\n".join(guide_lines))


def _safe_filename(text):
    """Convert text to a safe filename."""
    if not text:
        return "Untitled"
    # Remove characters that are problematic in filenames
    safe = re.sub(r'[<>:"/\\|?*]', '', text)
    # Truncate
    if len(safe) > 60:
        safe = safe[:60].rstrip()
    return safe


# ---------------------------------------------------------------------------
# Main: assemble everything
# ---------------------------------------------------------------------------

def generate(course_dir: Path):
    """Main entry point: generate brain/ and canvas_pages/ from inputs/."""
    inputs_dir = course_dir / "inputs"
    brain_dir = course_dir / "brain"
    canvas_dir = course_dir / "canvas_pages"

    # ── Find the Excel file ──
    excel_path = None
    for name in ["course_plan.xlsx", "course_planning.xlsx", "plan.xlsx"]:
        candidate = inputs_dir / name
        if candidate.exists():
            excel_path = candidate
            break
    if excel_path is None:
        # Try any xlsx file
        xlsx_files = list(inputs_dir.glob("*.xlsx"))
        xlsx_files = [f for f in xlsx_files if "roster" not in f.name.lower()]
        if xlsx_files:
            excel_path = xlsx_files[0]
    if excel_path is None:
        print("ERROR: No course planning Excel found in inputs/")
        print("  Expected: course_plan.xlsx (or any .xlsx that isn't a roster)")
        sys.exit(1)

    # ── Read config from Excel ──
    print(f"Reading course plan from: {excel_path.name}")
    config = read_excel_config(excel_path)
    print(f"  Course: {config['course_name']}")
    print(f"  Quarter: {config['quarter']}")
    print(f"  Sessions: {len(config.get('sessions', []))}")
    print(f"  Sections: {len(config['sections'])}")

    # ── Parse roster ──
    roster_path = None
    for name in ["roster.xlsx", "roster.xls"]:
        candidate = inputs_dir / name
        if candidate.exists():
            roster_path = candidate
            break
    if roster_path is None:
        roster_candidates = [f for f in inputs_dir.glob("*.xlsx") if "roster" in f.name.lower()]
        if roster_candidates:
            roster_path = roster_candidates[0]

    students = []
    if roster_path:
        students = parse_roster(roster_path, config["student_id_prefix"])
        print(f"  Students: {len(students)} (from {roster_path.name})")
    else:
        print("  Students: roster not found — skipping students.json")

    # ── Create directories ──
    brain_dir.mkdir(parents=True, exist_ok=True)
    (brain_dir / "class_logs").mkdir(exist_ok=True)

    # ── Generate brain/ files ──
    print(f"\nGenerating brain/ at {brain_dir}...")

    # Tracker JSON
    tracker = generate_tracker(config)
    with open(brain_dir / "course_tracker.json", "w", encoding="utf-8") as f:
        json.dump(tracker, f, indent=2, ensure_ascii=False)
    quiz_count = len(tracker.get("quiz_tracker", {}).get("quiz_schedule", []))
    print(f"  course_tracker.json     — {len(tracker['classes'])} classes, {quiz_count} quizzes detected")

    # ASSISTANT.md
    assistant_md = generate_assistant_md(config)
    with open(brain_dir / "ASSISTANT.md", "w", encoding="utf-8") as f:
        f.write(assistant_md)
    print(f"  ASSISTANT.md            — operating manual")

    # memory.md
    memory_md = generate_memory_md(config)
    with open(brain_dir / "memory.md", "w", encoding="utf-8") as f:
        f.write(memory_md)
    print(f"  memory.md               — rolling memory")

    # class_prep_reference.md
    prep_ref = generate_class_prep_reference(config)
    with open(brain_dir / "class_prep_reference.md", "w", encoding="utf-8") as f:
        f.write(prep_ref)
    print(f"  class_prep_reference.md — per-class detail")

    # SET protocol (with scores)
    set_protocol = generate_set_protocol(config)
    with open(brain_dir / "set_protocol.md", "w", encoding="utf-8") as f:
        f.write(set_protocol)
    print(f"  set_protocol.md         — with baseline scores")

    # Copy remaining shared protocols
    for sf in ["weekly_rhythm.md", "contingency.md", "skill_log.md", "ai_pedagogy.md"]:
        src = SHARED_DIR / sf
        if src.exists():
            shutil.copy2(src, brain_dir / sf)
            print(f"  {sf:<25s} — shared protocol")

    # Empty starter files
    for filename, title in [
        ("quiz_bank.md", "Quiz Bank"),
        ("exam_bank.md", "Exam Bank"),
        ("archive_completed_actions.md", "Archived Action Items"),
    ]:
        with open(brain_dir / filename, "w", encoding="utf-8") as f:
            f.write(f"# {title} — {config['course_code']}\n\n---\n")
        print(f"  {filename:<25s} — empty, ready to populate")

    # ── Generate students.json ──
    if students:
        students_path = course_dir / "students.json"
        students_json = {
            s["id"]: {"first_name": s["first_name"], "last_name": s["last_name"], "email": s["email"]}
            for s in students
        }
        with open(students_path, "w", encoding="utf-8") as f:
            json.dump(students_json, f, indent=2, ensure_ascii=False)
        print(f"  students.json           — {len(students)} students (GITIGNORE THIS)")

    # ── Generate Canvas pages ──
    print(f"\nGenerating canvas_pages/ at {canvas_dir}...")
    generate_canvas_pages(config, canvas_dir)
    session_files = list((canvas_dir / "Sessions").glob("*.md")) if (canvas_dir / "Sessions").exists() else []
    week_files = list((canvas_dir / "Weeks").glob("*.md")) if (canvas_dir / "Weeks").exists() else []
    print(f"  Session pages: {len(session_files)}")
    print(f"  Week pages:    {len(week_files)}")
    print(f"  Course Schedule + Setup Guide")

    # ── Generate .gitignore ──
    gitignore = "# FERPA\nstudents.json\ninputs/roster.*\n\n# Licensed content\ninputs/*.pdf\n\n# OS\n.DS_Store\n__pycache__/\n"
    with open(course_dir / ".gitignore", "w", encoding="utf-8") as f:
        f.write(gitignore)

    # ── Summary ──
    print(f"\n{'='*60}")
    print(f"  DONE: {config['course_code']} — {config['quarter']}")
    print(f"  brain/          → mount in your LLM assistant")
    print(f"  canvas_pages/   → copy to Canvas (LMS)")
    print(f"{'='*60}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate teaching assistant brain/ from a course planning Excel",
        epilog="Example: python generate_brain.py ../courses/MGMT162_Fall2026"
    )
    parser.add_argument("course_folder", type=Path,
                        help="Path to course folder (must contain inputs/ with course_plan.xlsx)")
    args = parser.parse_args()

    course_dir = args.course_folder.resolve()
    if not course_dir.exists():
        print(f"ERROR: Course folder not found: {course_dir}")
        sys.exit(1)
    if not (course_dir / "inputs").exists():
        print(f"ERROR: No inputs/ folder in {course_dir}")
        sys.exit(1)

    generate(course_dir)
